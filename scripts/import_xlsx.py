#!/usr/bin/env python3
"""Import Finance.xlsx into Solomon API via HTTP POST requests.

Usage:
    1. Start the API: docker compose up -d
    2. Create a user or set USER_ID env var
    3. Run: python3 scripts/import_xlsx.py ../Finance.xlsx

Requires: pip install openpyxl requests PyJWT
"""
import sys, os, re, math, datetime, json
import openpyxl
import requests
import jwt as pyjwt

API_URL = os.getenv("API_URL", "http://localhost:8080")
JWT_SECRET = os.getenv("JWT_SECRET", "dev-secret-change-in-prod")
USER_ID = int(os.getenv("USER_ID", "1"))

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "fev": 2, "feb": 2, "mar": 3, "apr": 4, "abr": 4,
    "mai": 5, "jun": 6, "jul": 7, "ago": 8, "aug": 8,
    "sep": 9, "set": 9, "out": 10, "oct": 10, "nov": 11, "dez": 12, "dec": 12,
}

def parse_sheet_date(name):
    """Extract year and month from sheet name like 'March 2021' or 'MAR 2026'."""
    parts = name.strip().lower().split()
    if len(parts) != 2:
        return None, None
    month_str, year_str = parts
    month = MONTH_MAP.get(month_str)
    try:
        year = int(year_str)
    except ValueError:
        return None, None
    return year, month

def to_cents(val):
    if val is None or val == '' or val == 0 or val == 0.0:
        return 0
    return int(round(float(val) * 100))

def parse_recurrence(val):
    if val is None:
        return 0
    s = str(val).strip().lower()
    if s == "fixa":
        return 1
    if s in ("variável", "variavel"):
        return 2
    return 0

def parse_status(val):
    if val is None:
        return 0
    s = str(val).strip().lower()
    if s == "confirmado":
        return 1
    return 0

def parse_installment(desc):
    """Extract installment from description like 'IPTU 3/12' or 'BYD 13/60'."""
    m = re.search(r'(\d+)/(\d+)', str(desc))
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None

def make_date(year, month, day_val):
    """Build a date string from year, month, and a day value (int or datetime)."""
    if isinstance(day_val, datetime.datetime):
        return day_val.strftime("%Y-%m-%d")
    if isinstance(day_val, datetime.date):
        return day_val.strftime("%Y-%m-%d")
    try:
        day = int(day_val)
        day = min(day, 28)
        return f"{year:04d}-{month:02d}-{day:02d}"
    except (ValueError, TypeError):
        return f"{year:04d}-{month:02d}-01"

def detect_format(ws, sheet_name):
    """Detect which format this sheet uses. Returns format identifier."""
    a1 = ws.cell(1, 1).value
    if a1 is None:
        return "unknown"
    a1s = str(a1).strip().lower()

    # Format D (ABR 2025+): A1 is status value (Confirmado/Não informado)
    if a1s in ("confirmado", "não informado"):
        return "D"

    # Format C (JAN-MAR 2025): A1 is "ENTRADAS", data starts row 2 or 3
    if a1s == "entradas":
        return "C"

    # Format B (Fev 2024 - Dez 2024): A1="Due date", 6 cols, values can be negative
    # Format A (Nov 2020 - Jan 2024): A1="Due date" or "Dia", 8 cols
    if a1s in ("due date", "dia"):
        b1 = ws.cell(1, 2).value
        if b1 and str(b1).strip().lower() == "type":
            return "B_with_type"
        if b1 and str(b1).strip().lower() in ("descrição", "descripton"):
            return "A_no_type"
        return "B_with_type"

    return "unknown"

def parse_format_A_no_type(ws, year, month):
    """Sep-Oct 2020: Dia | Descrição | Categoria | Valor | Status | ..."""
    txs = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]
        day, desc, cat, val, status = vals[0], vals[1], vals[2], vals[3], vals[4]
        if day is None or desc is None or val is None:
            continue
        if str(desc).upper() in ('TOTAL', 'SALDO'):
            continue
        cents = to_cents(abs(val))
        if cents == 0:
            continue
        tx_type = 1 if float(val) > 0 else 2
        inst_cur, inst_tot = parse_installment(desc)
        txs.append({
            "type": tx_type,
            "status": parse_status(status),
            "date": make_date(year, month, day),
            "description": str(desc).strip(),
            "category": str(cat).strip() if cat else "",
            "amount_cents": cents,
            "currency": "BRL",
            "recurrence_type": 0,
            "installment_cur": inst_cur,
            "installment_tot": inst_tot,
            "year_month": year * 100 + month,
        })
    return txs

def parse_format_B(ws, year, month):
    """Nov 2020 - Dez 2024: Due date | Type | Description | Category | Value | Status
       Value negative=expense, positive=income."""
    txs = []
    start_row = 2
    # Check if row 1 is header
    a1 = ws.cell(1, 1).value
    if a1 and str(a1).strip().lower() in ("due date", "dia"):
        start_row = 2
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row[:8]]
        date_val = vals[0]
        if date_val is None:
            continue
        # Determine column mapping
        if len(vals) >= 6:
            recur_val = vals[1]
            desc = vals[2]
            cat = vals[3]
            val = vals[4]
            status = vals[5]
        else:
            continue
        if desc is None or val is None:
            continue
        desc_s = str(desc).strip().upper()
        if desc_s in ('TOTAL', 'SALDO', 'DESCRIÇÃO', 'DESCRIPTON', 'TOTAL RECEITAS', 'TOTAL DESPESAS'):
            continue
        if str(cat).strip().upper() in ('RECEITAS', 'DESPESAS'):
            continue
        cents = to_cents(abs(val))
        if cents == 0:
            continue
        tx_type = 1 if float(val) > 0 else 2
        inst_cur, inst_tot = parse_installment(desc)
        txs.append({
            "type": tx_type,
            "status": parse_status(status),
            "date": make_date(year, month, date_val),
            "description": str(desc).strip(),
            "category": str(cat).strip() if cat else "",
            "amount_cents": cents,
            "currency": "BRL",
            "recurrence_type": parse_recurrence(recur_val),
            "installment_cur": inst_cur,
            "installment_tot": inst_tot,
            "year_month": year * 100 + month,
        })
    return txs

def parse_format_C(ws, year, month):
    """JAN-MAR 2025: ENTRADAS/SAÍDAS layout.
       Can be side-by-side (left=income cols A-F, right=expense cols H-M)
       or vertical (ENTRADAS section on top, SAÍDAS section below).
       Values always positive."""
    txs = []
    # First detect if vertical layout: look for SAÍDAS in column A
    saidas_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip().upper() in ('SAÍDAS', 'SAIDAS'):
            saidas_row = r
            break

    if saidas_row:
        # Vertical layout: income rows between header and SAÍDAS, expenses after SAÍDAS
        # Income section
        for row in ws.iter_rows(min_row=3, max_row=saidas_row - 1, values_only=False):
            vals = [c.value for c in row[:6]]
            date_val, recur, desc, cat, val, status = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5] if len(vals) > 5 else None
            if date_val is None or desc is None:
                continue
            if str(desc).strip().upper() in ('TOTAL', 'SALDO', 'DESCRIPTON'):
                continue
            if cat and str(cat).strip().upper() in ('RECEITAS', 'DESPESAS'):
                continue
            cents = to_cents(val)
            if cents > 0:
                inst_cur, inst_tot = parse_installment(desc)
                txs.append({
                    "type": 1, "status": parse_status(status),
                    "date": make_date(year, month, date_val),
                    "description": str(desc).strip(),
                    "category": str(cat).strip() if cat else "",
                    "amount_cents": cents, "currency": "BRL",
                    "recurrence_type": parse_recurrence(recur),
                    "installment_cur": inst_cur, "installment_tot": inst_tot,
                    "year_month": year * 100 + month,
                })
        # Expense section (skip SAÍDAS header + sub-header row)
        for row in ws.iter_rows(min_row=saidas_row + 2, max_row=ws.max_row, values_only=False):
            vals = [c.value for c in row[:6]]
            date_val, recur, desc, cat, val, status = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5] if len(vals) > 5 else None
            if date_val is None or desc is None:
                continue
            if str(desc).strip().upper() in ('TOTAL', 'SALDO', 'DESCRIPTON'):
                continue
            if cat and str(cat).strip().upper() in ('RECEITAS', 'DESPESAS'):
                continue
            cents = to_cents(val)
            if cents > 0:
                inst_cur, inst_tot = parse_installment(desc)
                txs.append({
                    "type": 2, "status": parse_status(status),
                    "date": make_date(year, month, date_val),
                    "description": str(desc).strip(),
                    "category": str(cat).strip() if cat else "",
                    "amount_cents": cents, "currency": "BRL",
                    "recurrence_type": parse_recurrence(recur),
                    "installment_cur": inst_cur, "installment_tot": inst_tot,
                    "year_month": year * 100 + month,
                })
    else:
        # Side-by-side layout: income left (A-F), expenses right (H-M)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
            vals = [c.value for c in row[:14]]
            # Left side: income (cols A-F)
            if vals[0] is not None and vals[2] is not None:
                date_val, recur, desc, cat, val, status = vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]
                if desc and str(desc).strip().upper() not in ('TOTAL', 'SALDO', 'DESCRIPTON'):
                    if cat and str(cat).strip().upper() not in ('RECEITAS', 'DESPESAS'):
                        cents = to_cents(val)
                        if cents > 0:
                            inst_cur, inst_tot = parse_installment(desc)
                            txs.append({
                                "type": 1, "status": parse_status(status),
                                "date": make_date(year, month, date_val),
                                "description": str(desc).strip(),
                                "category": str(cat).strip() if cat else "",
                                "amount_cents": cents, "currency": "BRL",
                                "recurrence_type": parse_recurrence(recur),
                                "installment_cur": inst_cur, "installment_tot": inst_tot,
                                "year_month": year * 100 + month,
                            })
            # Right side: expenses (cols H-M = index 7-12)
            if len(vals) > 12 and vals[7] is not None and vals[9] is not None:
                date_val, recur, desc, cat, val, status = vals[7], vals[8], vals[9], vals[10], vals[11], vals[12]
                if desc and str(desc).strip().upper() not in ('TOTAL', 'SALDO', 'DESCRIPTON'):
                    if cat and str(cat).strip().upper() not in ('RECEITAS', 'DESPESAS'):
                        cents = to_cents(val)
                        if cents > 0:
                            inst_cur, inst_tot = parse_installment(desc)
                            txs.append({
                                "type": 2, "status": parse_status(status),
                                "date": make_date(year, month, date_val),
                                "description": str(desc).strip(),
                                "category": str(cat).strip() if cat else "",
                                "amount_cents": cents, "currency": "BRL",
                                "recurrence_type": parse_recurrence(recur),
                                "installment_cur": inst_cur, "installment_tot": inst_tot,
                                "year_month": year * 100 + month,
                            })
    return txs

def parse_format_D(ws, year, month):
    """ABR 2025+: Status | Date | Type(opt) | Description | Category | Value
       Income rows before TOTAL RECEITAS, expenses after. Values always positive."""
    txs = []
    section = "income"  # starts as income
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row[:8]]
        # Detect TOTAL RECEITAS separator
        if vals[3] and str(vals[3]).strip().upper() == 'TOTAL' and vals[4] and str(vals[4]).strip().upper() == 'RECEITAS':
            section = "expense"
            continue
        if vals[3] and str(vals[3]).strip().upper() == 'TOTAL' and vals[4] and str(vals[4]).strip().upper() == 'DESPESAS':
            continue
        status_val = vals[0]
        date_val = vals[1]
        if status_val is None or date_val is None:
            continue
        if str(status_val).strip().lower() not in ("confirmado", "não informado"):
            continue
        recur = vals[2]
        desc = vals[3]
        cat = vals[4]
        val = vals[5]
        if desc is None:
            continue
        cents = to_cents(val)
        if cents == 0:
            continue
        tx_type = 1 if section == "income" else 2
        inst_cur, inst_tot = parse_installment(desc)
        txs.append({
            "type": tx_type,
            "status": parse_status(status_val),
            "date": make_date(year, month, date_val),
            "description": str(desc).strip(),
            "category": str(cat).strip() if cat else "",
            "amount_cents": cents,
            "currency": "BRL",
            "recurrence_type": parse_recurrence(recur),
            "installment_cur": inst_cur,
            "installment_tot": inst_tot,
            "year_month": year * 100 + month,
        })
    return txs

def generate_token():
    return pyjwt.encode(
        {"uid": USER_ID, "exp": datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=1)},
        JWT_SECRET, algorithm="HS256"
    )

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 import_xlsx.py <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    token = generate_token()
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    total_imported = 0
    total_skipped = 0
    total_errors = 0

    for sheet_name in wb.sheetnames:
        year, month = parse_sheet_date(sheet_name)
        if year is None or month is None:
            print(f"  SKIP sheet '{sheet_name}': cannot parse date")
            continue

        ws = wb[sheet_name]
        fmt = detect_format(ws, sheet_name)
        txs = []

        if fmt == "A_no_type":
            txs = parse_format_A_no_type(ws, year, month)
        elif fmt == "B_with_type":
            txs = parse_format_B(ws, year, month)
        elif fmt == "C":
            txs = parse_format_C(ws, year, month)
        elif fmt == "D":
            txs = parse_format_D(ws, year, month)
        else:
            print(f"  SKIP sheet '{sheet_name}': unknown format (A1={ws.cell(1,1).value})")
            total_skipped += 1
            continue

        imported = 0
        errors = 0
        for tx in txs:
            resp = requests.post(f"{API_URL}/v1/transactions", json=tx, headers=headers)
            if resp.status_code == 201:
                imported += 1
            else:
                errors += 1
                if errors <= 3:
                    print(f"    ERROR {resp.status_code}: {resp.text[:100]} | tx={tx['description']}")

        ym = year * 100 + month
        inc = sum(t["amount_cents"] for t in txs if t["type"] == 1)
        exp = sum(t["amount_cents"] for t in txs if t["type"] == 2)
        print(f"  {sheet_name:25s} ym={ym} fmt={fmt:15s} rows={len(txs):3d} imported={imported:3d} errors={errors} income=R${inc/100:,.2f} expense=R${exp/100:,.2f}")
        total_imported += imported
        total_errors += errors

    print(f"\nDone: {total_imported} imported, {total_errors} errors, {total_skipped} sheets skipped")

if __name__ == "__main__":
    main()
